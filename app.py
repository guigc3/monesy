import io
import os
import uuid
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from flask import Flask, g, jsonify, request, send_file, send_from_directory
from flask_cors import CORS

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:  # pragma: no cover
    pass

import auth
from db.repositories import (
    DEFAULT_SECOES,
    DEFAULT_SECOES_RECEITA,
    JsonRepository,
    _safe_write_json,
    get_repository,
)

_FRONTEND_DIST = os.path.join(os.path.dirname(os.path.abspath(__file__)), "frontend", "dist")
_STATIC_LEGACY = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static-legacy")

# Serve Vue dist se existir, senão serve o static legado
_static_folder = _FRONTEND_DIST if os.path.isdir(_FRONTEND_DIST) else _STATIC_LEGACY
app = Flask(__name__, static_folder=_static_folder, static_url_path="")
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DESIGN_SYSTEM_DIR = os.path.join(BASE_DIR, "Design System")
DATA_DIR = os.path.join(BASE_DIR, "data")
DATA_FILE = os.path.join(DATA_DIR, "gastos.json")
ASSINATURAS_FILE = os.path.join(DATA_DIR, "assinaturas.json")
FEATURES_FILE = os.path.join(DATA_DIR, "features.json")

MESES = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

os.makedirs(DATA_DIR, exist_ok=True)


def _ensure_design_system_static_link():
    """Espelha Design System/ em static-legacy/design-system (referência legado)."""
    static_ds = os.path.join(BASE_DIR, "static-legacy", "design-system")
    if os.path.exists(static_ds):
        return
    if not os.path.isdir(DESIGN_SYSTEM_DIR):
        return
    try:
        if os.name == "nt":
            import subprocess
            subprocess.run(
                ["cmd", "/c", "mklink", "/J", static_ds, DESIGN_SYSTEM_DIR],
                check=True,
                capture_output=True,
            )
        else:
            os.symlink(DESIGN_SYSTEM_DIR, static_ds, target_is_directory=True)
    except OSError:
        print(
            "AVISO: static-legacy/design-system ausente. Crie o link para Design System/ "
            "ou reinicie apos: mklink /J static-legacy\\design-system \"Design System\""
        )


_ensure_design_system_static_link()

auth.init_app(app)

# Registra teardown MySQL (noop em outros backends — g.pop retorna None)
from db.mysql_client import teardown_db as _mysql_teardown
app.teardown_appcontext(_mysql_teardown)


def _json_repo() -> JsonRepository:
    return JsonRepository(DATA_FILE, ASSINATURAS_FILE, FEATURES_FILE)


@app.before_request
def _ensure_repo_extension():
    app.extensions.setdefault("repositories", {})["json"] = _json_repo()


# ---------------------------------------------------------------------------
# Helpers (historico, normalizacao, agregacoes)
# ---------------------------------------------------------------------------


def _log_entry(acao, antes=None, depois=None):
    entry = {"ts": datetime.now().isoformat(timespec="seconds"), "acao": acao}
    if antes is not None:
        entry["antes"] = antes
    if depois is not None:
        entry["depois"] = depois
    return entry


def _snapshot(lanc):
    """Campos relevantes para o log de historico."""
    return {
        k: lanc.get(k)
        for k in ("descricao", "valor", "secao", "observacao", "tags", "pago", "investido", "tipo", "mes", "ano")
    }


def _assinatura_snapshot(item):
    return {
        k: item.get(k)
        for k in ("descricao", "data_inicio", "data_fim", "valor_mensal", "cartao")
    }


def resolve_descricao(_data, lanc):
    descricao = (lanc.get("descricao") or "").strip()
    return descricao or "Sem descrição"


def _descricao(lanc):
    descricao = (lanc.get("descricao") or "").strip()
    return descricao or "Sem descrição"


def normalize_tags(raw):
    if not raw:
        return []
    if isinstance(raw, str):
        raw = [t.strip() for t in raw.split(",")]
    seen = set()
    tags = []
    for tag in raw:
        name = (tag or "").strip()
        if not name:
            continue
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        tags.append(name)
    return tags


def collect_all_tags(lancamentos):
    seen = set()
    tags = []
    for lanc in lancamentos:
        for tag in normalize_tags(lanc.get("tags")):
            key = tag.lower()
            if key not in seen:
                seen.add(key)
                tags.append(tag)
    return sorted(tags, key=str.lower)


def filter_lancamentos(lancamentos, ano=None, mes=None, tipo=None, tag=None):
    """Helper preservado para o teste ``TestCalcTotais``."""
    result = lancamentos
    if ano is not None:
        result = [l for l in result if l.get("ano") == ano]
    if mes is not None:
        result = [l for l in result if l.get("mes") == mes]
    if tipo:
        result = [l for l in result if l.get("tipo") == tipo]
    if tag:
        tag_key = tag.strip().lower()
        result = [
            l
            for l in result
            if any(t.lower() == tag_key for t in normalize_tags(l.get("tags")))
        ]
    return result


def _totais_lancs(lancamentos):
    receitas = [l for l in lancamentos if l.get("tipo") == "receita"]
    despesas = [l for l in lancamentos if l.get("tipo") == "despesa"]
    entrada = sum(l.get("valor", 0) for l in receitas)
    entrada_investida = sum(l.get("valor", 0) for l in receitas if l.get("investido"))
    saida = sum(l.get("valor", 0) for l in despesas)
    saida_paga = sum(l.get("valor", 0) for l in despesas if l.get("pago"))
    saida_pendente = saida - saida_paga
    return {
        "entrada": round(entrada, 2),
        "entrada_investida": round(entrada_investida, 2),
        "saida": round(saida, 2),
        "saida_paga": round(saida_paga, 2),
        "saida_pendente": round(saida_pendente, 2),
        "caixa": round(entrada - entrada_investida - saida_paga, 2),
        "liquido": round(entrada - saida, 2),
    }


def calc_totais(data, ano, mes=None, tag=None):
    """Wrapper preservado para o teste ``TestCalcTotais`` (recebe ``data`` legado)."""
    return _totais_lancs(filter_lancamentos(data["lancamentos"], ano=ano, mes=mes, tag=tag))


def _build_sections(grouped, ordem_preferida):
    sections = []
    usadas = set()
    for secao in ordem_preferida:
        items = grouped.get(secao, [])
        if items:
            sections.append({
                "secao": secao,
                "itens": items,
                "total": round(sum(i["valor"] for i in items), 2),
            })
            usadas.add(secao)
    for secao, items in grouped.items():
        if secao in usadas:
            continue
        sections.append({
            "secao": secao,
            "itens": items,
            "total": round(sum(i["valor"] for i in items), 2),
        })
    return sections


def _group_lancs_into_secoes(secoes_despesa, secoes_receita, lancamentos):
    receitas_por_secao = defaultdict(list)
    despesas_por_secao = defaultdict(list)
    for lanc in lancamentos:
        item = {
            "id": lanc["id"],
            "descricao": _descricao(lanc),
            "valor": lanc.get("valor", 0),
            "observacao": lanc.get("observacao", ""),
            "secao": lanc.get("secao") or "Geral",
            "tags": normalize_tags(lanc.get("tags")),
            "pago": bool(lanc.get("pago", False)),
            "investido": bool(lanc.get("investido", False)),
            "criado_em": lanc.get("criado_em"),
            "ultima_alteracao": lanc.get("ultima_alteracao") or lanc.get("criado_em"),
        }
        if lanc.get("tipo") == "receita":
            receitas_por_secao[item["secao"]].append(item)
        else:
            despesas_por_secao[item["secao"]].append(item)
    receitas_sections = _build_sections(receitas_por_secao, secoes_receita)
    despesas_sections = _build_sections(despesas_por_secao, secoes_despesa)
    return receitas_sections, despesas_sections


def _secoes_key(tipo):
    if tipo == "receita":
        return "secoes_receita"
    if tipo == "despesa":
        return "secoes_despesa"
    return None


def _assinatura_ativa(item, ref_date=None):
    fim = item.get("data_fim")
    if not fim:
        return True
    ref = ref_date or datetime.now().strftime("%Y-%m-%d")
    return fim >= ref


def _enrich_assinatura(item):
    return {**item, "ativa": _assinatura_ativa(item)}


def safe_write_json(filepath, data):
    """Preservado apenas para compatibilidade com testes legados."""
    _safe_write_json(filepath, data)


# ---------------------------------------------------------------------------
# Rotas estaticas / meta
# ---------------------------------------------------------------------------


@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


@app.route("/design-system/<path:filename>")
def design_system(filename):
    return send_from_directory(DESIGN_SYSTEM_DIR, filename)


@app.route("/logos/<path:filename>")
def logos(filename):
    ds_logos = os.path.join(DESIGN_SYSTEM_DIR, "logos")
    if os.path.isfile(os.path.join(ds_logos, filename)):
        return send_from_directory(ds_logos, filename)
    return send_from_directory(os.path.join(BASE_DIR, "logos"), filename)


@app.route("/api/meta")
def meta():
    return jsonify({"meses": MESES})


@app.route("/api/config")
def public_config():
    backend = auth.storage_backend()
    if backend == "mysql":
        return jsonify({"backend": "mysql"})
    if backend == "supabase":
        return jsonify({
            "backend": "supabase",
            "supabase": {
                "url": os.environ.get("SUPABASE_URL", ""),
                "anon_key": os.environ.get("SUPABASE_ANON_KEY", ""),
            },
        })
    return jsonify({"backend": "json", "supabase": None})


# ---------------------------------------------------------------------------
# Auth MySQL (registro e login proprios)
# ---------------------------------------------------------------------------


@app.route("/api/auth/register", methods=["POST"])
def auth_register():
    if auth.storage_backend() != "mysql":
        return jsonify({"error": "Rota disponivel apenas no modo MySQL"}), 404
    from db.mysql_client import get_db
    from werkzeug.security import generate_password_hash

    body = request.get_json(silent=True) or {}
    email = (body.get("email") or "").strip().lower()
    password = body.get("password") or ""
    if not email or not password:
        return jsonify({"error": "Email e senha obrigatorios"}), 400
    if len(password) < 6:
        return jsonify({"error": "Senha deve ter pelo menos 6 caracteres"}), 400

    db = get_db()
    with db.cursor() as cur:
        cur.execute("SELECT id FROM users WHERE email = %s LIMIT 1", (email,))
        if cur.fetchone():
            return jsonify({"error": "Email ja cadastrado"}), 409
        user_id = str(uuid.uuid4())
        pw_hash = generate_password_hash(password)
        cur.execute(
            "INSERT INTO users (id, email, password_hash) VALUES (%s, %s, %s)",
            (user_id, email, pw_hash),
        )
    token = auth.create_token(user_id)
    return jsonify({"token": token, "user_id": user_id}), 201


@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    if auth.storage_backend() != "mysql":
        return jsonify({"error": "Rota disponivel apenas no modo MySQL"}), 404
    from db.mysql_client import get_db
    from werkzeug.security import check_password_hash

    body = request.get_json(silent=True) or {}
    email = (body.get("email") or "").strip().lower()
    password = body.get("password") or ""
    if not email or not password:
        return jsonify({"error": "Email e senha obrigatorios"}), 400

    db = get_db()
    with db.cursor() as cur:
        cur.execute(
            "SELECT id, password_hash FROM users WHERE email = %s LIMIT 1", (email,)
        )
        user = cur.fetchone()
    if not user or not check_password_hash(user["password_hash"], password):
        return jsonify({"error": "Email ou senha incorretos"}), 401
    token = auth.create_token(user["id"])
    return jsonify({"token": token, "user_id": user["id"]})


# ---------------------------------------------------------------------------
# Revisao mensal
# ---------------------------------------------------------------------------


@app.route("/api/revisao")
@auth.require_auth
def get_revisao():
    ano = request.args.get("ano", type=int)
    if not ano:
        return jsonify({"error": "Parametro 'ano' obrigatorio"}), 400
    revisados = get_repository().list_meses_revisados(g.user_id, ano)
    return jsonify({"ano": ano, "revisados": revisados})


@app.route("/api/revisao/marcar", methods=["POST"])
@auth.require_auth
def set_revisao():
    body = request.get_json(silent=True) or {}
    try:
        ano = int(body.get("ano"))
        mes = int(body.get("mes"))
    except (TypeError, ValueError):
        return jsonify({"error": "ano e mes obrigatorios"}), 400
    if mes < 1 or mes > 12:
        return jsonify({"error": "mes deve ser entre 1 e 12"}), 400
    if "revisado" not in body:
        return jsonify({"error": "Campo 'revisado' obrigatorio"}), 400

    revisados = get_repository().set_mes_revisado(g.user_id, ano, mes, bool(body["revisado"]))
    return jsonify({"ok": True, "ano": ano, "revisados": revisados})


# ---------------------------------------------------------------------------
# Anos
# ---------------------------------------------------------------------------


@app.route("/api/anos")
@auth.require_auth
def list_anos():
    anos = get_repository().list_anos(g.user_id)
    if not anos:
        anos = [datetime.now().year]
    return jsonify({"anos": anos})


@app.route("/api/anos", methods=["POST"])
@auth.require_auth
def create_ano():
    body = request.get_json(silent=True) or {}
    raw = body.get("ano")
    try:
        ano = int(raw)
    except (TypeError, ValueError):
        return jsonify({"error": "Ano invalido"}), 400
    if ano < 1900 or ano > 2200:
        return jsonify({"error": "Ano fora do intervalo permitido (1900-2200)"}), 400

    repo = get_repository()
    anos = repo.list_anos(g.user_id)
    if ano in anos:
        return jsonify({"error": "Ano ja cadastrado"}), 409

    repo.add_ano(g.user_id, ano)
    anos = sorted(set(anos) | {ano}, reverse=True)
    return jsonify({"ok": True, "ano": ano, "anos": anos}), 201


@app.route("/api/anos/<int:ano>", methods=["DELETE"])
@auth.require_auth
def delete_ano(ano):
    force = request.args.get("force", "").lower() in ("1", "true", "yes")
    repo = get_repository()
    qtd_lancs = repo.count_lancamentos_ano(g.user_id, ano)
    if qtd_lancs == 0 and not repo.has_ano(g.user_id, ano):
        return jsonify({"error": "Ano nao encontrado"}), 404
    if qtd_lancs > 0 and not force:
        return jsonify({
            "error": "Ano possui lancamentos",
            "lancamentos": qtd_lancs,
            "hint": "Envie ?force=true para excluir o ano e todos os lancamentos.",
        }), 409

    anos = repo.remove_ano(g.user_id, ano, cascade=qtd_lancs > 0)
    return jsonify({
        "ok": True,
        "ano": ano,
        "lancamentos_removidos": qtd_lancs,
        "anos": anos,
    })


# ---------------------------------------------------------------------------
# Secoes
# ---------------------------------------------------------------------------


@app.route("/api/secoes")
@auth.require_auth
def list_secoes():
    secoes_map = get_repository().list_secoes(g.user_id)
    return jsonify({
        "secoes": secoes_map["despesa"],
        "secoes_despesa": secoes_map["despesa"],
        "secoes_receita": secoes_map["receita"],
    })


@app.route("/api/secoes", methods=["POST"])
@auth.require_auth
def create_secao():
    body = request.get_json(silent=True) or {}
    tipo = (body.get("tipo") or "despesa").strip()
    nome = (body.get("nome") or "").strip()

    if tipo not in ("receita", "despesa"):
        return jsonify({"error": "Tipo deve ser 'receita' ou 'despesa'"}), 400
    if not nome:
        return jsonify({"error": "Nome da secao obrigatorio"}), 400

    repo = get_repository()
    secoes_map = repo.list_secoes(g.user_id)
    secoes = secoes_map["receita" if tipo == "receita" else "despesa"]
    if any(s.lower() == nome.lower() for s in secoes):
        return jsonify({"error": "Secao ja existe"}), 409

    repo.add_secao(g.user_id, tipo, nome)
    secoes = list(secoes) + [nome]
    return jsonify({"ok": True, "tipo": tipo, "nome": nome, "secoes": secoes}), 201


# ---------------------------------------------------------------------------
# Tags
# ---------------------------------------------------------------------------


@app.route("/api/tags")
@auth.require_auth
def list_tags():
    return jsonify({"tags": get_repository().list_tags(g.user_id)})


# ---------------------------------------------------------------------------
# Lancamentos
# ---------------------------------------------------------------------------


@app.route("/api/lancamentos")
@auth.require_auth
def list_lancamentos():
    ano = request.args.get("ano", type=int)
    mes = request.args.get("mes", type=int)
    tipo = request.args.get("tipo", "").strip() or None
    tag = request.args.get("tag", "").strip() or None
    if not ano:
        return jsonify({"error": "Parametro 'ano' obrigatorio"}), 400

    repo = get_repository()
    lancs = repo.list_lancamentos(g.user_id, ano, mes=mes, tipo=tipo)
    if tag:
        tag_key = tag.lower()
        lancs = [
            l for l in lancs
            if any(t.lower() == tag_key for t in normalize_tags(l.get("tags")))
        ]
    lancs.sort(key=lambda l: (l.get("mes", 0), l.get("tipo", ""), _descricao(l)))
    enriched = []
    for lanc in lancs:
        enriched.append({
            **lanc,
            "descricao": _descricao(lanc),
            "tags": normalize_tags(lanc.get("tags")),
        })
    return jsonify(enriched)


@app.route("/api/lancamentos", methods=["POST"])
@auth.require_auth
def create_lancamento():
    body = request.get_json(silent=True) or {}
    ano = body.get("ano")
    mes = body.get("mes")
    tipo = (body.get("tipo") or "").strip()
    valor = body.get("valor")
    descricao = (body.get("descricao") or body.get("categoria") or "").strip()
    secao = (body.get("secao") or "Geral").strip()
    observacao = (body.get("observacao") or "").strip()
    tags = normalize_tags(body.get("tags"))

    if not ano or not mes or tipo not in ("receita", "despesa"):
        return jsonify({"error": "Campos obrigatorios: ano, mes, tipo (receita|despesa)"}), 400
    try:
        valor = round(float(valor), 2)
    except (TypeError, ValueError):
        return jsonify({"error": "Valor invalido"}), 400
    if valor <= 0:
        return jsonify({"error": "Valor deve ser maior que zero"}), 400
    if not descricao:
        return jsonify({"error": "Informe a descricao"}), 400

    secao_final = secao or ("Receitas" if tipo == "receita" else "Geral")
    lanc = {
        "id": str(uuid.uuid4()),
        "ano": int(ano),
        "mes": int(mes),
        "tipo": tipo,
        "descricao": descricao,
        "secao": secao_final,
        "valor": valor,
        "observacao": observacao,
        "tags": tags,
        "criado_em": datetime.now().isoformat(timespec="seconds"),
    }
    history = _log_entry("criado", depois=_snapshot(lanc))
    inserted = get_repository().insert_lancamento(
        g.user_id, lanc, history_entry=history
    )
    return jsonify(inserted), 201


@app.route("/api/lancamentos/limpar-mes", methods=["DELETE"])
@auth.require_auth
def delete_mes():
    ano = request.args.get("ano", type=int)
    mes = request.args.get("mes", type=int)
    if not ano:
        return jsonify({"error": "Parametro 'ano' obrigatorio"}), 400
    if not mes or mes < 1 or mes > 12:
        return jsonify({"error": "Parametro 'mes' obrigatorio (1-12)"}), 400

    repo = get_repository()
    lancs = repo.list_lancamentos(g.user_id, ano, mes=mes, with_ultima_alteracao=False)
    if not lancs:
        return jsonify({"ok": True, "removidos": 0})
    now = datetime.now().isoformat(timespec="seconds")
    entries = [(l["id"], _log_entry("excluido", antes=_snapshot(l))) for l in lancs]
    n = repo.soft_delete_lancamentos_bulk(g.user_id, entries, excluido_em=now)
    return jsonify({"ok": True, "removidos": n})


@app.route("/api/lancamentos/<lanc_id>", methods=["PUT"])
@auth.require_auth
def update_lancamento(lanc_id):
    body = request.get_json(silent=True) or {}
    repo = get_repository()
    lanc = repo.get_lancamento(g.user_id, lanc_id)
    if not lanc:
        return jsonify({"error": "Lancamento nao encontrado"}), 404

    fields = {}
    antes = {}
    depois = {}

    def _track(campo, novo):
        velho = lanc.get(campo)
        if velho != novo:
            antes[campo] = velho
            depois[campo] = novo
            fields[campo] = novo

    if "valor" in body:
        try:
            v = round(float(body["valor"]), 2)
        except (TypeError, ValueError):
            return jsonify({"error": "Valor invalido"}), 400
        if v <= 0:
            return jsonify({"error": "Valor deve ser maior que zero"}), 400
        _track("valor", v)

    if "observacao" in body:
        _track("observacao", (body.get("observacao") or "").strip())

    if "tags" in body:
        _track("tags", normalize_tags(body.get("tags")))

    if "descricao" in body or "categoria" in body:
        descricao = (body.get("descricao") or body.get("categoria") or "").strip()
        if not descricao:
            return jsonify({"error": "Descricao obrigatoria"}), 400
        _track("descricao", descricao)

    ensure_secao_for = None
    if "secao" in body and lanc.get("tipo") in ("receita", "despesa"):
        fallback = "Receitas" if lanc.get("tipo") == "receita" else "Geral"
        nova_secao = (body.get("secao") or fallback).strip() or fallback
        _track("secao", nova_secao)
        ensure_secao_for = lanc.get("tipo")

    if "pago" in body and lanc.get("tipo") == "despesa":
        _track("pago", bool(body["pago"]))

    if "investido" in body:
        if lanc.get("tipo") != "receita":
            return jsonify({"error": "Investido aplica-se apenas a receitas"}), 400
        _track("investido", bool(body["investido"]))

    if not antes:
        # nada mudou — devolve o lancamento atual
        return jsonify({**lanc, "descricao": _descricao(lanc), "tags": normalize_tags(lanc.get("tags"))})

    keys = list(antes.keys())
    if keys == ["pago"] and depois["pago"]:
        acao = "pago"
    elif keys == ["pago"]:
        acao = "despago"
    elif keys == ["investido"] and depois["investido"]:
        acao = "investido"
    elif keys == ["investido"]:
        acao = "desinvestido"
    else:
        acao = "editado"
    history = _log_entry(acao, antes=antes, depois=depois)

    updated = repo.update_lancamento(
        g.user_id, lanc_id, fields,
        history_entry=history,
        ensure_secao_for=ensure_secao_for,
    )
    if updated is None:
        return jsonify({"error": "Lancamento nao encontrado"}), 404
    updated["descricao"] = _descricao(updated)
    updated["tags"] = normalize_tags(updated.get("tags"))
    return jsonify(updated)


@app.route("/api/lancamentos/<lanc_id>/historico")
@auth.require_auth
def get_historico(lanc_id):
    repo = get_repository()
    lanc = repo.get_lancamento(g.user_id, lanc_id)
    if not lanc:
        # pode estar na lixeira
        lanc = next((l for l in repo.list_lixeira(g.user_id) if l.get("id") == lanc_id), None)
        if not lanc:
            return jsonify({"error": "Lancamento nao encontrado"}), 404
    hist = repo.get_lancamento_historico(g.user_id, lanc_id)
    return jsonify({
        "id": lanc_id,
        "descricao": _descricao(lanc),
        "historico": list(reversed(hist)),
    })


@app.route("/api/lancamentos/<lanc_id>", methods=["DELETE"])
@auth.require_auth
def delete_lancamento(lanc_id):
    repo = get_repository()
    lanc = repo.get_lancamento(g.user_id, lanc_id)
    if not lanc:
        return jsonify({"error": "Lancamento nao encontrado"}), 404
    now = datetime.now().isoformat(timespec="seconds")
    history = _log_entry("excluido", antes=_snapshot(lanc))
    repo.soft_delete_lancamento(g.user_id, lanc_id, history, excluido_em=now)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Lixeira
# ---------------------------------------------------------------------------


@app.route("/api/lixeira")
@auth.require_auth
def list_lixeira():
    items = get_repository().list_lixeira(g.user_id)
    result = []
    for lanc in items:
        result.append({
            "id": lanc["id"],
            "ano": lanc.get("ano"),
            "mes": lanc.get("mes"),
            "mes_nome": MESES[lanc["mes"] - 1] if lanc.get("mes") and 1 <= lanc["mes"] <= 12 else "",
            "tipo": lanc.get("tipo"),
            "descricao": _descricao(lanc),
            "valor": lanc.get("valor"),
            "secao": lanc.get("secao"),
            "excluido_em": lanc.get("excluido_em"),
        })
    return jsonify({"lixeira": result, "total": len(result)})


@app.route("/api/lixeira/<lanc_id>/restaurar", methods=["POST"])
@auth.require_auth
def restaurar_lancamento(lanc_id):
    history = _log_entry("restaurado")
    lanc = get_repository().restore_lancamento(g.user_id, lanc_id, history)
    if not lanc:
        return jsonify({"error": "Lancamento nao encontrado na lixeira"}), 404
    return jsonify({"ok": True, "lancamento": lanc})


@app.route("/api/lixeira/<lanc_id>", methods=["DELETE"])
@auth.require_auth
def delete_permanente(lanc_id):
    if not get_repository().delete_permanente(g.user_id, lanc_id):
        return jsonify({"error": "Lancamento nao encontrado na lixeira"}), 404
    return jsonify({"ok": True})


@app.route("/api/lixeira", methods=["DELETE"])
@auth.require_auth
def esvaziar_lixeira():
    removidos = get_repository().empty_lixeira(g.user_id)
    return jsonify({"ok": True, "removidos": removidos})


# ---------------------------------------------------------------------------
# Resumo
# ---------------------------------------------------------------------------


@app.route("/api/resumo")
@auth.require_auth
def resumo():
    ano = request.args.get("ano", type=int)
    mes = request.args.get("mes", type=int)
    if not ano:
        return jsonify({"error": "Parametro 'ano' obrigatorio"}), 400

    tag = request.args.get("tag", "").strip() or None
    repo = get_repository()

    if mes:
        # resumo_mes() usa 1 chamada RPC (Supabase) ou 2 queries (JSON)
        # em vez das 3 chamadas separadas que existiam antes
        rdata = repo.resumo_mes(g.user_id, ano, mes)
        lancs = rdata.get("lancamentos") or []
        secoes_data = rdata.get("secoes") or {}
        secoes_despesa = secoes_data.get("despesa") or list(DEFAULT_SECOES)
        secoes_receita = secoes_data.get("receita") or list(DEFAULT_SECOES_RECEITA)
        if tag:
            tag_key = tag.lower()
            lancs = [
                l for l in lancs
                if any(t.lower() == tag_key for t in normalize_tags(l.get("tags")))
            ]
        receitas_secoes, despesas_secoes = _group_lancs_into_secoes(
            secoes_despesa, secoes_receita, lancs,
        )
        totais = _totais_lancs(lancs)
        receitas_flat = [item for s in receitas_secoes for item in s["itens"]]
        return jsonify({
            "ano": ano,
            "mes": mes,
            "mes_nome": MESES[mes - 1] if 1 <= mes <= 12 else "",
            "totais": totais,
            "receitas": receitas_flat,
            "receitas_por_secao": receitas_secoes,
            "despesas_por_secao": despesas_secoes,
        })

    # Modo grafico: agregacoes do ano em uma unica query, processamento em Python
    lancs_ano = repo.list_lancamentos(g.user_id, ano, with_ultima_alteracao=False)
    by_mes = defaultdict(list)
    for l in lancs_ano:
        by_mes[l.get("mes")].append(l)
    mensal = []
    for m in range(1, 13):
        totais = _totais_lancs(by_mes.get(m, []))
        mensal.append({"mes": m, "mes_nome": MESES[m - 1], **totais})

    totais_ano = _totais_lancs(lancs_ano)
    por_descricao = defaultdict(lambda: {"receita": 0.0, "despesa": 0.0})
    for lanc in lancs_ano:
        nome = _descricao(lanc)
        por_descricao[nome][lanc["tipo"]] += lanc.get("valor", 0)

    ranking_despesas = sorted(
        [
            {"descricao": nome, "total": round(vals["despesa"], 2)}
            for nome, vals in por_descricao.items()
            if vals["despesa"] > 0
        ],
        key=lambda x: x["total"],
        reverse=True,
    )[:10]

    return jsonify({
        "ano": ano,
        "totais_ano": totais_ano,
        "mensal": mensal,
        "ranking_despesas": ranking_despesas,
    })


# ---------------------------------------------------------------------------
# Excel template / import
# ---------------------------------------------------------------------------

TEMPLATE_HEADERS = [
    "Ano", "Mes", "Tipo", "Descricao", "Secao", "Valor", "Observacao", "Tags",
]

TEMPLATE_EXEMPLOS = [
    [2026, 1, "receita", "Salário",      "Receitas",       5000.00, "",            "salario,mensal"],
    [2026, 1, "despesa", "Condomínio",   "Despesas fixas",  850.00, "vencimento dia 10", "casa,fixo"],
    [2026, 1, "despesa", "Cartão Nubank","Cartões",        1200.00, "",            "cartao"],
]

MESES_MAP = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "março": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8, "setembro": 9,
    "outubro": 10, "novembro": 11, "dezembro": 12,
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}


def _parse_mes(raw):
    if raw is None or raw == "":
        return None
    if isinstance(raw, int):
        return raw if 1 <= raw <= 12 else None
    try:
        n = int(str(raw).strip())
        if 1 <= n <= 12:
            return n
    except ValueError:
        pass
    key = str(raw).strip().lower()
    return MESES_MAP.get(key)


def _parse_tipo(raw):
    if not raw:
        return None
    key = str(raw).strip().lower()
    if key in ("receita", "entrada", "r"):
        return "receita"
    if key in ("despesa", "saida", "saída", "d"):
        return "despesa"
    return None


def _parse_valor(raw):
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        return round(float(raw), 2)
    text = str(raw).strip().replace("R$", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return round(float(text), 2)
    except ValueError:
        return None


@app.route("/api/template-excel")
@auth.require_auth
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lancamentos"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="D49B3B", end_color="D49B3B", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col, name in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, exemplo in enumerate(TEMPLATE_EXEMPLOS, 2):
        for col_idx, val in enumerate(exemplo, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    larguras = [8, 6, 12, 28, 22, 12, 30, 22]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Instrucoes")
    instrucoes = [
        ["Coluna",      "Obrigatorio", "Formato / Exemplo"],
        ["Ano",         "Sim",         "Numero inteiro (ex.: 2026)"],
        ["Mes",         "Sim",         "Numero 1-12 ou nome do mes (ex.: Janeiro)"],
        ["Tipo",        "Sim",         "receita ou despesa"],
        ["Descricao",   "Sim",         "Texto livre (ex.: Condominio)"],
        ["Secao",       "Nao",         "Despesas fixas, Cartoes, Receitas, etc."],
        ["Valor",       "Sim",         "Numero positivo (ex.: 850,00)"],
        ["Observacao",  "Nao",         "Texto livre"],
        ["Tags",        "Nao",         "Separadas por virgula (ex.: casa,fixo)"],
    ]
    for r, linha in enumerate(instrucoes, 1):
        for c, val in enumerate(linha, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="modelo-gastos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/lancamentos/import-excel", methods=["POST"])
@auth.require_auth
def import_excel():
    if "arquivo" not in request.files:
        return jsonify({"error": "Envie o arquivo no campo 'arquivo'"}), 400
    f = request.files["arquivo"]
    if not f.filename:
        return jsonify({"error": "Arquivo vazio"}), 400
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Use um arquivo .xlsx"}), 400

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
    except Exception as exc:
        return jsonify({"error": f"Nao foi possivel abrir o arquivo: {exc}"}), 400

    sheet_name = request.form.get("aba") or "Lancamentos"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"error": "Planilha vazia"}), 400

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

    def col(name):
        candidatos = {
            "ano":        ["ano"],
            "mes":        ["mes", "mês"],
            "tipo":       ["tipo"],
            "descricao":  ["descricao", "descrição", "descricao do lancamento"],
            "secao":      ["secao", "seção", "categoria"],
            "valor":      ["valor", "valor (r$)", "preco"],
            "observacao": ["observacao", "observação", "obs"],
            "tags":       ["tags", "etiquetas"],
        }
        for alvo in candidatos[name]:
            if alvo in header:
                return header.index(alvo)
        return -1

    idx = {k: col(k) for k in
           ["ano", "mes", "tipo", "descricao", "secao", "valor", "observacao", "tags"]}
    obrigatorios = ["ano", "mes", "tipo", "descricao", "valor"]
    faltando = [c for c in obrigatorios if idx[c] < 0]
    if faltando:
        return jsonify({
            "error": f"Cabecalho invalido. Colunas obrigatorias ausentes: {', '.join(faltando)}",
            "esperado": TEMPLATE_HEADERS,
        }), 400

    novos = []
    erros = []
    agora = datetime.now().isoformat(timespec="seconds")

    def get_row(row, key):
        i = idx[key]
        return row[i] if i >= 0 and i < len(row) else None

    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(c in (None, "") for c in row):
            continue
        try:
            ano = int(get_row(row, "ano")) if get_row(row, "ano") not in (None, "") else None
        except (TypeError, ValueError):
            ano = None
        mes = _parse_mes(get_row(row, "mes"))
        tipo = _parse_tipo(get_row(row, "tipo"))
        descricao = (str(get_row(row, "descricao")).strip() if get_row(row, "descricao") else "")
        secao = (str(get_row(row, "secao")).strip() if get_row(row, "secao") else "")
        valor = _parse_valor(get_row(row, "valor"))
        observacao = (str(get_row(row, "observacao")).strip() if get_row(row, "observacao") else "")
        tags_raw = get_row(row, "tags")

        problemas = []
        if not ano or ano < 1900 or ano > 2200:
            problemas.append("Ano invalido")
        if not mes:
            problemas.append("Mes invalido")
        if not tipo:
            problemas.append("Tipo deve ser 'receita' ou 'despesa'")
        if not descricao:
            problemas.append("Descricao obrigatoria")
        if valor is None or valor <= 0:
            problemas.append("Valor invalido")
        if problemas:
            erros.append({"linha": row_num, "problemas": problemas})
            continue

        if not secao:
            secao = "Receitas" if tipo == "receita" else "Geral"

        tags = normalize_tags(tags_raw if isinstance(tags_raw, list) else str(tags_raw or ""))
        novos.append({
            "id": str(uuid.uuid4()),
            "ano": int(ano),
            "mes": int(mes),
            "tipo": tipo,
            "descricao": descricao,
            "secao": secao,
            "valor": valor,
            "observacao": observacao,
            "tags": tags,
            "criado_em": agora,
        })

    criados = get_repository().bulk_insert_lancamentos(g.user_id, novos)
    return jsonify({
        "ok": True,
        "criados": criados,
        "erros": erros,
        "total_linhas": len(rows) - 1,
    })


# ---------------------------------------------------------------------------
# Assinaturas
# ---------------------------------------------------------------------------


def _parse_iso_date(value, field_name, required=False):
    if value is None or (isinstance(value, str) and not value.strip()):
        if required:
            raise ValueError(f"Campo '{field_name}' obrigatorio")
        return None
    s = str(value).strip()[:10]
    try:
        datetime.strptime(s, "%Y-%m-%d")
    except ValueError:
        raise ValueError(f"Data invalida em '{field_name}' (use AAAA-MM-DD)")
    return s


def _validar_datas_assinatura(data_inicio, data_fim):
    inicio = _parse_iso_date(data_inicio, "data_inicio", required=True)
    fim = _parse_iso_date(data_fim, "data_fim", required=False)
    if fim and fim < inicio:
        raise ValueError("data_fim nao pode ser anterior a data_inicio")
    return inicio, fim


@app.route("/api/assinaturas/cartoes")
@auth.require_auth
def list_cartoes_assinaturas():
    return jsonify({"cartoes": get_repository().list_cartoes(g.user_id)})


@app.route("/api/assinaturas")
@auth.require_auth
def list_assinaturas():
    cartao = (request.args.get("cartao") or "").strip() or None
    apenas_ativas = request.args.get("ativas", "").lower() in ("1", "true", "sim")
    items = get_repository().list_assinaturas(g.user_id)
    if cartao:
        items = [a for a in items if (a.get("cartao") or "").lower() == cartao.lower()]
    items.sort(key=lambda a: ((a.get("cartao") or "").lower(), (a.get("descricao") or "").lower()))
    enriched = [_enrich_assinatura(a) for a in items]
    if apenas_ativas:
        enriched = [a for a in enriched if a["ativa"]]
    total_mensal_ativas = round(
        sum(float(a.get("valor_mensal") or 0) for a in enriched if a.get("ativa")), 2
    )
    return jsonify({"assinaturas": enriched, "total_mensal_ativas": total_mensal_ativas})


@app.route("/api/assinaturas", methods=["POST"])
@auth.require_auth
def create_assinatura():
    body = request.get_json(silent=True) or {}
    descricao = (body.get("descricao") or "").strip()
    cartao = (body.get("cartao") or "").strip()
    if not descricao:
        return jsonify({"error": "Informe a descricao"}), 400
    if not cartao:
        return jsonify({"error": "Informe o cartao de credito"}), 400
    try:
        data_inicio, data_fim = _validar_datas_assinatura(
            body.get("data_inicio"), body.get("data_fim")
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    try:
        valor_mensal = round(float(body.get("valor_mensal")), 2)
    except (TypeError, ValueError):
        return jsonify({"error": "Valor mensal invalido"}), 400
    if valor_mensal <= 0:
        return jsonify({"error": "Valor mensal deve ser maior que zero"}), 400

    item = {
        "id": str(uuid.uuid4()),
        "descricao": descricao,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "valor_mensal": valor_mensal,
        "cartao": cartao,
        "criado_em": datetime.now().isoformat(timespec="seconds"),
    }
    history = _log_entry("criado", depois=_assinatura_snapshot(item))
    inserted = get_repository().insert_assinatura(g.user_id, item, history_entry=history)
    return jsonify(_enrich_assinatura(inserted)), 201


@app.route("/api/assinaturas/<item_id>", methods=["PUT"])
@auth.require_auth
def update_assinatura(item_id):
    body = request.get_json(silent=True) or {}
    repo = get_repository()
    item = repo.get_assinatura(g.user_id, item_id)
    if not item:
        return jsonify({"error": "Assinatura nao encontrada"}), 404

    fields = {}
    antes = {}
    depois = {}

    def _track(campo, novo):
        velho = item.get(campo)
        if velho != novo:
            antes[campo] = velho
            depois[campo] = novo
            fields[campo] = novo

    if "descricao" in body:
        descricao = (body.get("descricao") or "").strip()
        if not descricao:
            return jsonify({"error": "Informe a descricao"}), 400
        _track("descricao", descricao)

    if "cartao" in body:
        cartao = (body.get("cartao") or "").strip()
        if not cartao:
            return jsonify({"error": "Informe o cartao de credito"}), 400
        _track("cartao", cartao)

    if "valor_mensal" in body:
        try:
            valor_mensal = round(float(body["valor_mensal"]), 2)
        except (TypeError, ValueError):
            return jsonify({"error": "Valor mensal invalido"}), 400
        if valor_mensal <= 0:
            return jsonify({"error": "Valor mensal deve ser maior que zero"}), 400
        _track("valor_mensal", valor_mensal)

    if "data_inicio" in body or "data_fim" in body:
        try:
            data_inicio, data_fim = _validar_datas_assinatura(
                body.get("data_inicio", item.get("data_inicio")),
                body.get("data_fim", item.get("data_fim")),
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        _track("data_inicio", data_inicio)
        _track("data_fim", data_fim)

    history = _log_entry("editado", antes=antes, depois=depois) if antes else None
    updated = repo.update_assinatura(g.user_id, item_id, fields, history_entry=history)
    if updated is None:
        return jsonify({"error": "Assinatura nao encontrada"}), 404
    return jsonify(_enrich_assinatura(updated))


@app.route("/api/assinaturas/<item_id>/historico")
@auth.require_auth
def get_assinatura_historico(item_id):
    repo = get_repository()
    item = repo.get_assinatura(g.user_id, item_id)
    if not item:
        return jsonify({"error": "Assinatura nao encontrada"}), 404
    hist = repo.get_assinatura_historico(g.user_id, item_id)
    return jsonify({
        "id": item_id,
        "descricao": item.get("descricao") or "",
        "historico": list(reversed(hist)),
    })


@app.route("/api/assinaturas/<item_id>", methods=["DELETE"])
@auth.require_auth
def delete_assinatura(item_id):
    if not get_repository().delete_assinatura(g.user_id, item_id):
        return jsonify({"error": "Assinatura nao encontrada"}), 404
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Features (changelog global)
# ---------------------------------------------------------------------------


@app.route("/api/features")
def list_features():
    # Em modo supabase/mysql exige token; em modo json e publico.
    if auth.storage_backend() in ("supabase", "mysql"):
        token = request.headers.get("Authorization", "")
        if not token.lower().startswith("bearer "):
            return jsonify({"error": "Token de autenticacao ausente"}), 401
    repo = get_repository()
    items = repo.list_features()
    items.sort(key=lambda x: x.get("implementado_em") or "", reverse=True)
    return jsonify({"features": items, "total": len(items)})


# ── SPA catch-all ────────────────────────────────────────────────────────────
# Flask's built-in static serving intercepts requests first.
# We use @errorhandler(404) to catch files-not-found and serve index.html
# for Vue Router URLs (anything that's not /api/*)
@app.errorhandler(404)
def spa_fallback(e):
    """Para qualquer 404 que não seja /api/, serve o Vue SPA (index.html)."""
    dist = app.static_folder
    if not dist:
        return jsonify({"error": "Not found"}), 404

    # Deixa APIs retornarem 404 normalmente
    from flask import request as _req
    if _req.path.startswith("/api/"):
        return jsonify({"error": "Not found"}), 404

    index = os.path.join(dist, "index.html")
    if os.path.exists(index):
        return send_from_directory(dist, "index.html")

    return jsonify({"error": "Not found"}), 404


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5001"))
    backend = auth.storage_backend()
    print("=" * 60)
    print("  Monesy")
    print(f"  Backend: {backend}")
    if backend == "mysql":
        import db.mysql_client as _mc
        host = _mc._env("MYSQL_HOST")
        db_port = _mc._env("MYSQL_PORT", "3306")
        db_name = _mc._env("MYSQL_DATABASE")
        print(f"  MySQL: {host}:{db_port}/{db_name}")
    elif backend == "json":
        print(f"  Dados: {DATA_FILE}")
        print(f"  Assinaturas: {ASSINATURAS_FILE}")
        print(f"  Features: {FEATURES_FILE}")
    else:
        print(f"  Supabase: {os.environ.get('SUPABASE_URL', '?')}")
    print(f"  Servidor: http://localhost:{port}")
    print("=" * 60)
    app.run(host="0.0.0.0", port=port, debug=False)
