"""Importa dados da planilha Planilha gastos.xlsx para gastos/data/gastos.json."""

import json
import os
import uuid
from datetime import datetime

import openpyxl

MESES_PT = {
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARÇO": 3,
    "MARCO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
    "DEZEMBRO": 12,
}

DEFAULT_XLSX = os.path.join(os.path.expanduser("~"), "Downloads", "Planilha gastos.xlsx")
DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "gastos.json")


def _normalize_month(name):
    if not name:
        return None
    key = str(name).strip().upper()
    key = key.replace("Ç", "C").replace("Ã", "A")
    return MESES_PT.get(key)


def _month_starts(ws):
    blocks = []
    for col in range(1, ws.max_column + 1):
        label = ws.cell(1, col).value
        mes = _normalize_month(label)
        if mes:
            blocks.append((col, mes, str(label).strip()))
    return blocks


def _extract_entries(ws, start_col, mes_num):
    cat_col = start_col
    val_col = start_col + 1
    alt_col = start_col + 2
    entries = []
    section = "Geral"
    mode = None

    for row in range(2, 60):
        label = ws.cell(row, cat_col).value
        if not label:
            continue
        text = str(label).strip()
        val = ws.cell(row, val_col).value
        if val is None:
            val = ws.cell(row, alt_col).value
        if val is not None:
            try:
                val = float(val)
            except (TypeError, ValueError):
                val = None

        if text == "ENTRADA" and row < 12:
            mode = "receita"
            section = "Receitas"
            continue
        if text == "SAIDA" and row < 55:
            mode = "despesa"
            if section == "Receitas":
                section = "Despesas fixas"
            continue
        if text == "OUTROS":
            section = "Outros"
            continue
        if text in ("TOTAL", "LIQUIDO"):
            continue
        if text == "ENTRADA" and row >= 55:
            continue
        if text == "SAIDA" and row >= 55:
            continue

        if mode in ("receita", "despesa") and val not in (None, 0):
            entries.append(
                {
                    "tipo": mode,
                    "secao": section,
                    "descricao": text,
                    "valor": round(val, 2),
                    "mes": mes_num,
                }
            )

    return entries


def import_year(wb_path, year_sheet="2026"):
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    if year_sheet not in wb.sheetnames:
        raise ValueError(f"Aba '{year_sheet}' nao encontrada")
    ws = wb[year_sheet]
    year = int(year_sheet) if year_sheet.isdigit() else datetime.now().year

    secoes = set()
    lancamentos = []

    for start_col, mes_num, _ in _month_starts(ws):
        for item in _extract_entries(ws, start_col, mes_num):
            if item["tipo"] == "despesa":
                secoes.add(item["secao"])
            lancamentos.append(
                {
                    "id": str(uuid.uuid4()),
                    "ano": year,
                    "mes": item["mes"],
                    "tipo": item["tipo"],
                    "descricao": item["descricao"],
                    "secao": item["secao"],
                    "valor": item["valor"],
                    "observacao": "",
                    "criado_em": datetime.now().isoformat(timespec="seconds"),
                }
            )

    return {
        "secoes_despesa": sorted(secoes),
        "lancamentos": lancamentos,
    }


def main():
    xlsx = os.environ.get("PLANILHA_GASTOS", DEFAULT_XLSX)
    sheet = os.environ.get("PLANILHA_ANO", "2026")
    if not os.path.exists(xlsx):
        print(f"Arquivo nao encontrado: {xlsx}")
        return 1
    data = import_year(xlsx, sheet)
    os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Importados {len(data['lancamentos'])} lancamentos para {DATA_FILE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
